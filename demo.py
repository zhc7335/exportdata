# coding=utf-8
import cx_Oracle
from openpyxl import Workbook


# 连接oracle，读取数据
def link_Oracle(wb):
    create_wb(wb, 'helloDatas.xlsx')
    ws = wb.create_sheet(title='我是第一个title', index=0)
    # 用户名、密码、"ip/库名"
    connection = cx_Oracle.connect("root", "126315", "10.7.137.76/DNF")
    cursor = connection.cursor()
    sql = '''
            select * from zhc.dept 
          '''
    cursor.execute(sql)
    # row = cursor.fetchone()
    # print(row)
    list_A_B = []
    # 遍历游标
    rowcount = 1
    for a in cursor:
        rowcount = rowcount + 1
        list_A_B.append(a)

    # 读取表字段值
    db_title = [i[0] for i in cursor.description]
    # 遍历表字段值
    for i, description in enumerate(db_title):
        ws.cell(row=1, column=1 + i).value = description

    # 读取数据到excel
    for rowNum in range(1, rowcount):
        A = ws.cell(row=rowNum, column=1)
        B = ws.cell(row=rowNum, column=2)
        # list_A_B索引从0开始，所以-1，list中存的是tuple，后面[0]获取的是tuple中的对应元素
        A.value = list_A_B[rowNum - 1][0]
        B.value = list_A_B[rowNum - 1][1]

    cursor.close()
    connection.close()
    wb.save('helloDatas.xlsx')


# 创建Excel
def create_wb(wb, filename):
    wb.save(filename=filename)
    print("新建Excel：" + filename + "成功")


if __name__ == '__main__':
    wb = Workbook()
    link_Oracle(wb)
